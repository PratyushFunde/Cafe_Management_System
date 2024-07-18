from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

n=int(input("Enter number of student : "))

students={}
# {'1': {'name': 'a', 'Maths': 2, 'english': 2, 'science': 2}, '2': {'name': 'b', 'Maths': 3, 'english': 3, 'science': 3}}

for i in range(0,n):
    n=input("Enter Roll No. : ").strip()
    name=input("Enter name of student :")
    math=int(input("Enter marks of Maths"))
    science=int(input("Enter marks of science"))
    english=int(input("Enter marks of English"))
    students[n]={}
    students[n]["roll_no"]=n
    students[n]['name']=name
    students[n]['Maths']=math
    students[n]['english']=english
    students[n]['science']=english
 


wb=Workbook()

wb.title=("Test")

ws=wb.create_sheet("Records")


headings=['Roll_No','Name','Math','Science','English']
ws.append(headings)

for person in students:
    grades=list(students[person].values())
    ws.append(grades)
    wb.save("Test.xlsx")


